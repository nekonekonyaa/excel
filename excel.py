
import openpyxl
book= openpyxl.load_workbook('./no1.xlsx')
sheet= book['Sheet1']

#最終行を取得
max=sheet.max_row
for i in range(max):
    #pytho3.8以上でないと以下のif文は使えない
    if (link :=sheet.cell(row=i+1,column=2).hyperlink) != None:
        print(f'行{i+1}:{link.target}')
    else:
        print(f'行{i+1}:no hyperlink')
    # print(sheet.cell(row=i+1,column=2).hyperlink)
    # print()

